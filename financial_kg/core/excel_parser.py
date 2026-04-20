"""Excel解析器 - 提取单元格、公式、合并单元格，识别子表边界"""
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from typing import Optional
import datetime

from ..models.cell import CellNode
from ..models.graph import KnowledgeGraph


class ExcelParser:
    """解析Excel工作簿，构建知识图谱"""

    def __init__(self, file_path: str, target_sheets: Optional[list] = None):
        self.file_path = file_path
        self.target_sheets = target_sheets
        self.wb = None
        self.wb_values = None  # data_only=True 用于获取计算值

    def load(self) -> None:
        """加载Excel文件"""
        self.wb = openpyxl.load_workbook(self.file_path, data_only=False)
        self.wb_values = openpyxl.load_workbook(self.file_path, data_only=True)

    def parse(self, graph_name: str = "financial_model") -> KnowledgeGraph:
        """主入口：解析所有目标sheet，构建知识图谱"""
        if self.wb is None:
            self.load()

        # 确定要解析的sheet
        sheets_to_parse = self.target_sheets or list(self.wb.sheetnames)

        # 创建图谱
        graph = KnowledgeGraph(
            name=graph_name,
            source_file=self.file_path,
            sheets=sheets_to_parse
        )

        for sheet_name in sheets_to_parse:
            if sheet_name not in self.wb.sheetnames:
                continue
            ws = self.wb[sheet_name]
            ws_values = self.wb_values[sheet_name]
            self._parse_sheet(ws, ws_values, sheet_name, graph)

        return graph

    def _parse_sheet(self, ws, ws_values, sheet_name: str, graph: KnowledgeGraph) -> None:
        """解析单个sheet"""
        # 1. 获取合并单元格信息
        merged_info = self._extract_merged_cells(ws)

        # 2. 提取所有非空单元格
        for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
            for cell in row:
                if cell.value is None:
                    continue

                # 创建节点
                node = self._create_cell_node(
                    cell, ws, ws_values, sheet_name, merged_info
                )
                graph.add_node(node)

    def _extract_merged_cells(self, ws) -> dict:
        """提取合并单元格信息，返回 {左上角地址: 合并范围字符串}"""
        merged_info = {}
        for merged_range in ws.merged_cells.ranges:
            top_left = merged_range.start_cell.coordinate
            merged_info[top_left] = str(merged_range)
        return merged_info

    def _create_cell_node(self, cell, ws, ws_values, sheet_name: str, merged_info: dict) -> CellNode:
        """从Excel单元格创建CellNode"""
        col_letter = get_column_letter(cell.column)
        row_num = cell.row
        cell_id = CellNode.make_id(sheet_name, row_num, col_letter)
        address = f"{sheet_name}!{col_letter}{row_num}"

        # 原始值和公式
        raw_value = cell.value
        formula_raw = None
        if isinstance(raw_value, str) and raw_value.startswith("="):
            formula_raw = raw_value
            raw_value = None  # 公式单元格的原始值为None

        # 尝试获取Excel计算值
        computed_value = None
        try:
            value_cell = ws_values.cell(row=row_num, column=cell.column)
            computed_value = value_cell.value
        except Exception:
            pass

        # 确定值类型
        value_type = self._determine_value_type(raw_value or computed_value)

        # 判断是否合并单元格
        is_merged = cell.coordinate in merged_info
        merge_range = merged_info.get(cell.coordinate)

        # 判断是否表头（简单规则：前3行或包含特定关键词）
        is_header = row_num <= 3 or self._is_header_content(raw_value)

        # 获取行类别（来自合并单元格）
        row_category = self._get_row_category(ws, row_num, merged_info)

        return CellNode(
            id=cell_id,
            sheet=sheet_name,
            row=row_num,
            col=col_letter,
            col_index=cell.column - 1,
            address=address,
            value=self._serialize_value(raw_value or computed_value),
            value_type=value_type,
            formula_raw=formula_raw,
            computed_value=computed_value,
            is_header=is_header,
            is_merged=is_merged,
            merge_range=merge_range,
            row_category=row_category,
            parse_status="raw",
        )

    def _determine_value_type(self, value) -> str:
        """确定值的类型"""
        if value is None:
            return "empty"
        if isinstance(value, (int, float)):
            return "number"
        if isinstance(value, str):
            return "string"
        if isinstance(value, datetime.datetime):
            return "date"
        if isinstance(value, bool):
            return "bool"
        return "unknown"

    def _serialize_value(self, value):
        """序列化特殊值类型"""
        if isinstance(value, datetime.datetime):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, datetime.date):
            return value.strftime("%Y-%m-%d")
        return value

    def _is_header_content(self, value) -> bool:
        """判断是否表头内容"""
        if value is None:
            return False
        header_keywords = ["类别", "序号", "参数", "项目", "合计", "单位", "年份", "备注"]
        if isinstance(value, str):
            return any(kw in value for kw in header_keywords)
        return False

    def _get_row_category(self, ws, row_num: int, merged_info: dict) -> Optional[str]:
        """获取行类别（来自B列的合并单元格）"""
        # 查找B列的合并单元格中包含当前行的范围
        b_col_letter = "B"
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_col == 2 and merged_range.min_row <= row_num <= merged_range.max_row:
                # 这是B列的合并范围，获取左上角的值作为类别
                top_left_cell = ws.cell(row=merged_range.min_row, column=2)
                if top_left_cell.value and isinstance(top_left_cell.value, str):
                    return top_left_cell.value
        return None


def col_to_index(col: str) -> int:
    """列字母转数字索引（A=0, B=1, ..., Z=25, AA=26）"""
    return column_index_from_string(col) - 1


def index_to_col(index: int) -> str:
    """数字索引转列字母"""
    return get_column_letter(index + 1)